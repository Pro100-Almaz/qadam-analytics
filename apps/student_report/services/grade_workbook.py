"""
Renders collected grade data (see grade_sheet.py) into a styled .xlsx workbook.

Two layouts, chosen by the `ordering` query parameter:

`subject` (default)
    One sheet per subject. Rows are students, columns are that subject's
    assignments in date order.

`student`
    One sheet per student. Each sheet stacks one small table per subject: a
    header row of that subject's assignments, and a single row of grades
    labelled with the subject name. Subjects do not share assignments, so they
    cannot share a column grid — hence many small tables rather than one.

In both layouts a cell holds the grade, or "-" when none was given. A grade
carrying a teacher comment is filled amber and gets an Excel note with the
comment text, so the grade itself stays a number that Excel can still sum.
"""

import re

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MISSING_MARK = '-'
COMMENT_AUTHOR = 'Qadam Analytics'

NAME_COLUMN_WIDTH = 34
GRADE_COLUMN_WIDTH = 18
HEADER_ROW_HEIGHT = 46

_TITLE_FONT = Font(bold=True, size=14, color='1F3864')
_SUBTITLE_FONT = Font(size=10, color='5A6B84')
_NOTE_FONT = Font(size=9, italic=True, color='7A879A')
_HEADER_FONT = Font(bold=True, size=10, color='FFFFFF')
_LABEL_FONT = Font(bold=True, size=10, color='1F3864')
_MISSING_FONT = Font(size=11, italic=True, color='9AA5B1')

_HEADER_FILL = PatternFill('solid', fgColor='1F3864')
_LABEL_FILL = PatternFill('solid', fgColor='D9E2F3')
_BAND_FILL = PatternFill('solid', fgColor='F4F7FB')
_COMMENT_FILL = PatternFill('solid', fgColor='FFF2CC')

_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
_NAME_ALIGN = Alignment(horizontal='left', vertical='center', indent=1)

_THIN = Side(style='thin', color='C8D2E0')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')

_LEGEND = (
    f'"{MISSING_MARK}" — no grade recorded. '
    'Amber cells carry a teacher comment; hover over one to read it.'
)


# ── Sheet naming ──

def _sheet_title(name, used):
    """
    An Excel-legal, unique sheet name.

    Excel rejects \\ / * ? : [ ] and anything past 31 characters, and refuses
    two sheets with the same name — so collisions get a numeric suffix that is
    trimmed back into the limit.
    """
    clean = _INVALID_SHEET_CHARS.sub(' ', name).strip() or 'Sheet'
    candidate = clean[:31]
    counter = 2
    while candidate.lower() in used:
        suffix = f' ({counter})'
        candidate = f'{clean[:31 - len(suffix)]}{suffix}'
        counter += 1
    used.add(candidate.lower())
    return candidate


# ── Cell writers ──

def _write_page_header(ws, title, subtitle, width):
    """Title and period block above every table. Returns the next free row."""
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = _SUBTITLE_FONT
    if width > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.row_dimensions[1].height = 22
    return 4


def _write_column_headers(ws, row, columns, corner_label):
    """The corner cell plus one header per assignment: title over its date."""
    corner = ws.cell(row=row, column=1, value=corner_label)
    corner.font = _HEADER_FONT
    corner.fill = _HEADER_FILL
    corner.alignment = _HEADER_ALIGN
    corner.border = _BORDER

    for index, column in enumerate(columns, start=2):
        cell = ws.cell(
            row=row, column=index,
            value=f'{column.title}\n{column.date:%d.%m.%Y}',
        )
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER
        note = Comment(
            f'{column.title}\n'
            f'Date: {column.date:%d.%m.%Y}\n'
            f'Category: {column.category}\n'
            f'Max grade: {column.max_grade}',
            COMMENT_AUTHOR,
        )
        note.width, note.height = 220, 100
        cell.comment = note

    ws.row_dimensions[row].height = HEADER_ROW_HEIGHT


def _write_label(ws, row, value, banded=False):
    cell = ws.cell(row=row, column=1, value=value)
    cell.font = _LABEL_FONT
    cell.alignment = _NAME_ALIGN
    cell.border = _BORDER
    cell.fill = _LABEL_FILL if not banded else _BAND_FILL
    return cell


def _write_grade(ws, row, column_index, grade_cell, banded=False):
    cell = ws.cell(row=row, column=column_index)
    cell.alignment = _CENTER_ALIGN
    cell.border = _BORDER

    if grade_cell is None or grade_cell.grade is None:
        cell.value = MISSING_MARK
        cell.font = _MISSING_FONT
    else:
        cell.value = grade_cell.grade

    if grade_cell is not None and grade_cell.comment:
        cell.fill = _COMMENT_FILL
        note = Comment(grade_cell.comment, COMMENT_AUTHOR)
        note.width, note.height = 280, 130
        cell.comment = note
    elif banded:
        cell.fill = _BAND_FILL

    return cell


def _write_legend(ws, row, width):
    cell = ws.cell(row=row, column=1, value=_LEGEND)
    cell.font = _NOTE_FONT
    if width > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)


def _set_widths(ws, column_count):
    ws.column_dimensions['A'].width = NAME_COLUMN_WIDTH
    for index in range(2, column_count + 2):
        ws.column_dimensions[get_column_letter(index)].width = GRADE_COLUMN_WIDTH


def _empty_note(ws, row, message):
    cell = ws.cell(row=row, column=1, value=message)
    cell.font = _NOTE_FONT
    return row + 1


# ── Layouts ──

def _period_subtitle(data):
    return (
        f'Quarter {data.quarter} · '
        f'{data.period_start:%d.%m.%Y} — {data.period_end:%d.%m.%Y} · '
        f'{data.academic_year_label}'
    )


def _render_by_subject(wb, data):
    """One sheet per subject: students down, assignments across."""
    used_titles = set()

    for subject in data.subjects:
        ws = wb.create_sheet(_sheet_title(subject.subject_name, used_titles))
        width = len(subject.columns) + 1
        _set_widths(ws, len(subject.columns))

        row = _write_page_header(
            ws,
            f'{subject.subject_name} — {data.class_group_label}',
            _period_subtitle(data),
            width,
        )

        if not subject.columns:
            _empty_note(ws, row, 'No assignments in this quarter.')
            continue

        _write_column_headers(ws, row, subject.columns, 'Student')
        header_row = row
        row += 1

        if not data.students:
            _empty_note(ws, row, 'No students enrolled in this class group.')
            continue

        for offset, student in enumerate(data.students):
            banded = offset % 2 == 1
            _write_label(ws, row, student.name, banded=banded)
            for index, column in enumerate(subject.columns, start=2):
                _write_grade(
                    ws, row, index,
                    data.cell(student.student_id, column.assignment_id),
                    banded=banded,
                )
            row += 1

        _write_legend(ws, row + 1, width)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=2)


def _render_by_student(wb, data):
    """One sheet per student, stacking a small table per subject."""
    used_titles = set()
    width = data.widest_subject + 1

    for student in data.students:
        ws = wb.create_sheet(_sheet_title(student.name, used_titles))
        _set_widths(ws, data.widest_subject)

        row = _write_page_header(
            ws,
            f'{student.name} — {data.class_group_label}',
            _period_subtitle(data),
            width,
        )
        first_table_row = row

        if not data.subjects:
            _empty_note(ws, row, 'No subjects offered to this class group.')
            continue

        for subject in data.subjects:
            if not subject.columns:
                _write_label(ws, row, subject.subject_name)
                note = ws.cell(
                    row=row, column=2, value='No assignments in this quarter.',
                )
                note.font = _NOTE_FONT
                note.alignment = _NAME_ALIGN
                row += 2
                continue

            _write_column_headers(ws, row, subject.columns, 'Subject')
            row += 1

            _write_label(ws, row, subject.subject_name)
            for index, column in enumerate(subject.columns, start=2):
                _write_grade(
                    ws, row, index,
                    data.cell(student.student_id, column.assignment_id),
                )
            row += 2

        _write_legend(ws, row, width)
        ws.freeze_panes = ws.cell(row=first_table_row, column=2)


# ── Entry point ──

def build_grade_workbook(data, ordering='subject') -> Workbook:
    """
    Render `data` (a GradeSheetData) into a workbook.

    A workbook with no sheets is invalid, so an empty result still gets one
    sheet saying so rather than producing a file Excel refuses to open.
    """
    wb = Workbook()
    wb.remove(wb.active)

    if ordering == 'student':
        _render_by_student(wb, data)
    else:
        _render_by_subject(wb, data)

    if not wb.sheetnames:
        ws = wb.create_sheet('No data')
        ws.column_dimensions['A'].width = 60
        _write_page_header(
            ws,
            f'No grades for {data.class_group_label}',
            _period_subtitle(data),
            1,
        )
        _empty_note(
            ws, 4,
            'Nothing was recorded for this class group in this quarter.',
        )

    return wb
