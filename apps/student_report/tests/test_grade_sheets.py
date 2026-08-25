import io
from datetime import date

import pytest
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status

from apps.home.models import (
    HomeroomTeacherAssignment,
    SubjectAssignment,
    SubjectGrade,
)
from core.factories import (
    AdminUserFactory,
    ClassGroupFactory,
    EnrollmentFactory,
    ParentFactory,
    StudentFactory,
    SubjectFactory,
    SubjectOfferingFactory,
    TeacherFactory,
    TeachingAssignmentFactory,
)

XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)


def open_workbook(response):
    return load_workbook(io.BytesIO(response.content))


@pytest.fixture
def year_with_quarters(academic_year):
    """The active year with Q1/Q2 date ranges, which assignments are filed by."""
    academic_year.q1_start = date(2025, 9, 1)
    academic_year.q1_end = date(2025, 10, 31)
    academic_year.q2_start = date(2025, 11, 1)
    academic_year.q2_end = date(2025, 12, 31)
    academic_year.save()
    return academic_year


@pytest.fixture
def sheet_setup(year_with_quarters):
    """
    One class group, two students, two subjects, graded work in Q1 and Q2.

    Math has two Q1 assignments (one grade carries a comment, one student is
    ungraded on the second); Art has one. A Q2 Math assignment exists so that
    quarter filtering has something to exclude.
    """
    class_group = ClassGroupFactory(academic_year=year_with_quarters)

    alice = StudentFactory(academic_year=year_with_quarters)
    alice.user.first_name, alice.user.last_name = 'Alice', 'Abbott'
    alice.user.save()
    bob = StudentFactory(academic_year=year_with_quarters)
    bob.user.first_name, bob.user.last_name = 'Bob', 'Barnes'
    bob.user.save()

    for student in (alice, bob):
        EnrollmentFactory(
            student=student,
            class_group=class_group,
            academic_year=year_with_quarters,
        )

    math = SubjectFactory(name='Mathematics')
    art = SubjectFactory(name='Art')
    math_offering = SubjectOfferingFactory(
        subject=math, class_group=class_group, academic_year=year_with_quarters,
    )
    art_offering = SubjectOfferingFactory(
        subject=art, class_group=class_group, academic_year=year_with_quarters,
    )

    # Created out of date order on purpose — columns must come back sorted.
    quiz = SubjectAssignment.objects.create(
        offering=math_offering, title='Quiz 2', max_grade=10,
        date=date(2025, 10, 20), category='lesson',
    )
    entry_test = SubjectAssignment.objects.create(
        offering=math_offering, title='Entry test', max_grade=10,
        date=date(2025, 9, 10), category='exam',
    )
    q2_exam = SubjectAssignment.objects.create(
        offering=math_offering, title='Winter exam', max_grade=10,
        date=date(2025, 12, 1), category='exam',
    )
    still_life = SubjectAssignment.objects.create(
        offering=art_offering, title='Still life', max_grade=5,
        date=date(2025, 9, 25), category='lesson',
    )

    SubjectGrade.objects.create(
        assignment=entry_test, student=alice, grade=9, comments='Great start',
    )
    SubjectGrade.objects.create(assignment=entry_test, student=bob, grade=6)
    SubjectGrade.objects.create(assignment=quiz, student=alice, grade=8)
    SubjectGrade.objects.create(assignment=still_life, student=alice, grade=5)
    SubjectGrade.objects.create(assignment=q2_exam, student=alice, grade=4)

    homeroom_teacher = TeacherFactory()
    HomeroomTeacherAssignment.objects.create(
        teacher=homeroom_teacher,
        class_group=class_group,
        academic_year=year_with_quarters,
    )
    math_teacher = TeacherFactory()
    TeachingAssignmentFactory(teacher=math_teacher, offering=math_offering)

    return {
        'academic_year': year_with_quarters,
        'class_group': class_group,
        'alice': alice,
        'bob': bob,
        'math': math,
        'art': art,
        'entry_test': entry_test,
        'quiz': quiz,
        'q2_exam': q2_exam,
        'homeroom_teacher': homeroom_teacher,
        'math_teacher': math_teacher,
    }


def class_group_url(class_group):
    return reverse(
        'student-report-api:class-group-grade-sheet',
        kwargs={'class_group_id': class_group.pk},
    )


def student_url(student):
    return reverse(
        'student-report-api:student-grade-sheet',
        kwargs={'student_id': student.pk},
    )


@pytest.mark.django_db
class TestClassGroupGradeSheet:

    def test_homeroom_teacher_gets_a_sheet_per_subject(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']), {'quarter': 1},
        )

        assert response.status_code == status.HTTP_200_OK
        assert response['Content-Type'] == XLSX_CONTENT_TYPE

        wb = open_workbook(response)
        assert set(wb.sheetnames) == {'Art', 'Mathematics'}

    def test_columns_are_date_ordered_and_quarter_scoped(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']), {'quarter': 1},
        )

        ws = open_workbook(response)['Mathematics']
        headers = [ws.cell(row=4, column=col).value for col in range(1, 5)]

        # Entry test (10.09) precedes Quiz 2 (20.10); the Q2 exam is absent.
        assert headers[0] == 'Student'
        assert headers[1].startswith('Entry test')
        assert headers[2].startswith('Quiz 2')
        assert headers[3] is None

    def test_cells_carry_grades_dashes_and_comments(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']), {'quarter': 1},
        )

        ws = open_workbook(response)['Mathematics']
        # Row 5 is Abbott, row 6 is Barnes — ordered by last name.
        assert ws.cell(row=5, column=1).value == 'Alice Abbott'
        assert ws.cell(row=5, column=2).value == 9
        assert ws.cell(row=5, column=2).comment.text.endswith('Great start')
        assert ws.cell(row=6, column=1).value == 'Bob Barnes'
        assert ws.cell(row=6, column=2).value == 6
        assert ws.cell(row=6, column=3).value == '-'

    def test_subject_filter_leaves_one_sheet(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']),
            {'quarter': 1, 'subject': sheet_setup['math'].pk},
        )

        assert response.status_code == status.HTTP_200_OK
        assert open_workbook(response).sheetnames == ['Mathematics']

    def test_student_ordering_gives_a_sheet_per_student(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']),
            {'quarter': 1, 'ordering': 'student'},
        )

        wb = open_workbook(response)
        assert set(wb.sheetnames) == {'Alice Abbott', 'Bob Barnes'}

        ws = wb['Alice Abbott']
        labels = [
            ws.cell(row=row, column=1).value for row in range(4, 12)
        ]
        assert 'Art' in labels
        assert 'Mathematics' in labels

    def test_subject_teacher_needs_the_subject_parameter(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['math_teacher'].user)
        url = class_group_url(sheet_setup['class_group'])

        denied = client.get(url, {'quarter': 1})
        assert denied.status_code == status.HTTP_403_FORBIDDEN

        allowed = client.get(url, {'quarter': 1, 'subject': sheet_setup['math'].pk})
        assert allowed.status_code == status.HTTP_200_OK

    def test_subject_teacher_cannot_reach_a_subject_they_do_not_teach(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['math_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']),
            {'quarter': 1, 'subject': sheet_setup['art'].pk},
        )
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_admin_reads_any_class_group(self, authenticated_client, sheet_setup):
        client = authenticated_client(AdminUserFactory())
        response = client.get(
            class_group_url(sheet_setup['class_group']), {'quarter': 1},
        )
        assert response.status_code == status.HTTP_200_OK

    def test_unrelated_student_is_denied(self, authenticated_client, sheet_setup):
        client = authenticated_client(sheet_setup['alice'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']), {'quarter': 1},
        )
        assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
class TestStudentGradeSheet:

    def test_student_reads_their_own_sheet(self, authenticated_client, sheet_setup):
        client = authenticated_client(sheet_setup['alice'].user)
        response = client.get(
            student_url(sheet_setup['alice']),
            {'quarter': 1, 'ordering': 'student'},
        )

        assert response.status_code == status.HTTP_200_OK
        wb = open_workbook(response)
        assert wb.sheetnames == ['Alice Abbott']

    def test_sheet_holds_only_that_student(self, authenticated_client, sheet_setup):
        client = authenticated_client(sheet_setup['alice'].user)
        response = client.get(student_url(sheet_setup['alice']), {'quarter': 1})

        ws = open_workbook(response)['Mathematics']
        assert ws.cell(row=5, column=1).value == 'Alice Abbott'
        assert ws.cell(row=6, column=1).value is None

    def test_parent_reads_their_child(self, authenticated_client, sheet_setup):
        parent = ParentFactory()
        parent.students.add(sheet_setup['alice'])

        client = authenticated_client(parent.user)
        response = client.get(student_url(sheet_setup['alice']), {'quarter': 1})
        assert response.status_code == status.HTTP_200_OK

    def test_parent_of_another_child_is_denied(
        self, authenticated_client, sheet_setup,
    ):
        parent = ParentFactory()
        parent.students.add(sheet_setup['bob'])

        client = authenticated_client(parent.user)
        response = client.get(student_url(sheet_setup['alice']), {'quarter': 1})
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_classmate_is_denied(self, authenticated_client, sheet_setup):
        client = authenticated_client(sheet_setup['bob'].user)
        response = client.get(student_url(sheet_setup['alice']), {'quarter': 1})
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_homeroom_teacher_reads_the_student(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(student_url(sheet_setup['alice']), {'quarter': 1})
        assert response.status_code == status.HTTP_200_OK

    def test_subject_teacher_needs_the_subject_parameter(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['math_teacher'].user)
        url = student_url(sheet_setup['alice'])

        assert client.get(url, {'quarter': 1}).status_code == status.HTTP_403_FORBIDDEN
        allowed = client.get(url, {'quarter': 1, 'subject': sheet_setup['math'].pk})
        assert allowed.status_code == status.HTTP_200_OK

    def test_unenrolled_student_reports_the_missing_enrollment(
        self, authenticated_client, sheet_setup,
    ):
        loner = StudentFactory(academic_year=sheet_setup['academic_year'])
        client = authenticated_client(loner.user)
        response = client.get(student_url(loner), {'quarter': 1})

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert 'student' in response.json()

    def test_outsider_gets_403_not_the_enrollment_detail(
        self, authenticated_client, sheet_setup,
    ):
        loner = StudentFactory(academic_year=sheet_setup['academic_year'])
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(student_url(loner), {'quarter': 1})

        assert response.status_code == status.HTTP_403_FORBIDDEN


@pytest.mark.django_db
class TestGradeSheetParameters:

    def test_quarter_is_required(self, authenticated_client, sheet_setup):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(class_group_url(sheet_setup['class_group']))

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert 'quarter' in response.json()

    @pytest.mark.parametrize('quarter', ['0', '5', 'two', ''])
    def test_quarter_is_validated(self, authenticated_client, sheet_setup, quarter):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']), {'quarter': quarter},
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_ordering_and_subject_are_mutually_exclusive(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']),
            {
                'quarter': 1,
                'subject': sheet_setup['math'].pk,
                'ordering': 'student',
            },
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert 'ordering' in response.json()

    def test_unknown_ordering_is_rejected(self, authenticated_client, sheet_setup):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']),
            {'quarter': 1, 'ordering': 'teacher'},
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_subject_not_taught_to_the_class_is_rejected(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(AdminUserFactory())
        other_subject = SubjectFactory(name='Astronomy')
        response = client.get(
            class_group_url(sheet_setup['class_group']),
            {'quarter': 1, 'subject': other_subject.pk},
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert 'subject' in response.json()

    def test_unconfigured_quarter_dates_are_reported(
        self, authenticated_client, sheet_setup,
    ):
        client = authenticated_client(sheet_setup['homeroom_teacher'].user)
        response = client.get(
            class_group_url(sheet_setup['class_group']), {'quarter': 3},
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert 'quarter' in response.json()

    def test_filename_survives_non_ascii_names(
        self, authenticated_client, sheet_setup,
    ):
        student = sheet_setup['alice']
        student.user.first_name, student.user.last_name = 'Айгүл', 'Сериккызы'
        student.user.save()

        client = authenticated_client(student.user)
        response = client.get(student_url(student), {'quarter': 1})

        disposition = response['Content-Disposition']
        assert response.status_code == status.HTTP_200_OK
        assert "filename*=UTF-8''" in disposition
        assert disposition.startswith('attachment; filename="Grades_')
